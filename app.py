
from flask import Flask, render_template, request, redirect, session, send_file, flash, jsonify
import psycopg2
import psycopg2.extras
import os
from psycopg2 import pool
from datetime import datetime
from zoneinfo import ZoneInfo
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from werkzeug.security import generate_password_hash, check_password_hash

DATABASE_URL = os.environ.get("DATABASE_URL")

try:
    db_pool = psycopg2.pool.SimpleConnectionPool(
        1,
        10,
        DATABASE_URL,
        sslmode="require"
    )
    print("Pool de conexões criado com sucesso")
except Exception as e:
    print("ERRO AO CRIAR POOL:", e)
    db_pool = None

app = Flask(__name__)
app.secret_key = "quermesse_secret"

# =========================
# CONEXÃO POSTGRES
# =========================
def conectar():
    try:
        return db_pool.getconn()
    except Exception as e:
        print("ERRO AO OBTER CONEXÃO:", e)
        return None

def fechar_conexao(conn):
    try:
        db_pool.putconn(conn)
    except Exception as e:
        print("ERRO AO DEVOLVER CONEXÃO:", e)

def agora_amazonas():
    return datetime.now(ZoneInfo("America/Manaus"))
    
# =========================
# LOGIN
# =========================
@app.route("/")
def login():
    return render_template("login.html")

@app.route("/autenticar", methods=["GET", "POST"])
def autenticar():

    # Se alguém tentar acessar via GET, volta para login
    if request.method == "GET":
        return redirect("/")

    usuario = request.form["usuario"]
    senha = request.form["senha"]

    conn = conectar()
    if not conn:
        return "Erro ao conectar no banco", 500

    c = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    c.execute("SELECT * FROM usuarios WHERE usuario=%s", (usuario,))
    user = c.fetchone()
    conn.close()

    if user and check_password_hash(user["senha"], senha):
        session["usuario_id"] = user["id"]
        session["usuario"] = user["usuario"]  # pode manter se quiser
        session["perfil"] = user.get("perfil", "usuario")
        return redirect("/dashboard")

    flash("Usuário ou senha inválidos", "danger")
    return redirect("/")

@app.route("/logout")
def logout():
    session.clear()
    return redirect("/")
    
# =========================
# CADASTRO USUÁRIO
# =========================
@app.route("/cadastro", methods=["GET", "POST"])
def cadastro():
    if not session.get("usuario"):
        return redirect("/")

    conn = conectar()
    c = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    if request.method == "POST":
        nome_usuario = request.form["nome_usuario"]
        usuario = request.form["usuario"]
        senha = generate_password_hash(request.form["senha"])
        perfil = request.form["perfil"]

        # Verifica se já existe
        c.execute("SELECT id FROM usuarios WHERE usuario = %s", (usuario,))
        if c.fetchone():
            conn.close()
            flash("Usuário já cadastrado!", "danger")
            return redirect("/cadastro")

        c.execute(
            "INSERT INTO usuarios (nome_usuario,usuario, senha, perfil) VALUES (%s, %s, %s, %s)",
            (nome_usuario, usuario, senha, perfil)
        )
        conn.commit()
        flash("Usuário cadastrado com sucesso!", "success")
        conn.close()
        return redirect("/cadastro")

    # 🔹 IMPORTANTE: SEMPRE EXECUTA NO GET
    c.execute("SELECT id,nome_usuario, usuario, perfil FROM usuarios ORDER BY usuario ASC")
    usuarios = c.fetchall()
    conn.close()

    return render_template("cadastro.html", usuarios=usuarios)


# =========================
# EDITAR USUÁRIO
# =========================    
@app.route("/editar_usuario/<int:id>", methods=["GET","POST"])
def editar_usuario(id):
    if not session.get("usuario"):
        return redirect("/")

    conn = conectar()
    c = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    # Busca usuário
    c.execute("SELECT id, nome_usuario, usuario, perfil FROM usuarios WHERE id = %s", (id,))
    usuario = c.fetchone()

    if not usuario:
        conn.close()
        flash("Usuário não encontrado.", "danger")
        return redirect("/cadastro")

    if request.method == "POST":
        novo_nome_usuario = request.form["nome_usuario"]
        novo_usuario = request.form["usuario"]
        novo_perfil = request.form["perfil"]
        nova_senha = request.form["senha"]
    
        if nova_senha:
            senha_hash = generate_password_hash(nova_senha)
            c.execute(
                "UPDATE usuarios SET nome_usuario = %s, usuario = %s, perfil = %s, senha = %s WHERE id = %s",
                (novo_nome_usuario, novo_usuario, novo_perfil, senha_hash, id)
            )
        else:
            c.execute(
                "UPDATE usuarios SET nome_usuario = %s, usuario = %s, perfil = %s WHERE id = %s",
                (novo_nome_usuario, novo_usuario, novo_perfil, id)
            )
    
        conn.commit()
        conn.close()
    
        flash("Usuário atualizado com sucesso!", "success")
        return redirect("/cadastro")

    conn.close()
    return render_template("editar_usuario.html", usuario=usuario)

# =========================
# EXCLUIR USUÁRIO
# =========================
@app.route("/excluir_usuario/<int:id>", methods=["POST"])
def excluir_usuario(id):
    if not session.get("usuario"):
        return redirect("/")

    conn = conectar()
    c = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    # Busca usuário a excluir
    c.execute("SELECT usuario FROM usuarios WHERE id = %s", (id,))
    usuario_excluir = c.fetchone()

    if not usuario_excluir:
        conn.close()
        flash("Usuário não encontrado.", "danger")
        return redirect("/cadastro")

    # 🔒 NÃO permite excluir o próprio usuário logado
    if usuario_excluir["usuario"] == session["usuario"]:
        conn.close()
        flash("Você não pode excluir o próprio usuário!", "danger")
        return redirect("/cadastro")

    # Exclui
    c.execute("DELETE FROM usuarios WHERE id = %s", (id,))
    conn.commit()
    conn.close()

    flash("Usuário excluído com sucesso!", "success")
    return redirect("/cadastro")

# =========================
# DASHBOARD
# =========================
@app.route("/dashboard")
def dashboard():
    if "usuario" not in session:
        return redirect("/")
    return render_template("dashboard.html")

# =========================
# PRODUTOS
# =========================
@app.route("/produtos", methods=["GET", "POST"])
def produtos():
    if "usuario" not in session:
        return redirect("/")

    if session.get("perfil") != "administrador":
        return redirect("/dashboard")

    conn = conectar()
    c = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    if request.method == "POST":
        descricao = request.form["descricao"]
        valor = float(request.form["valor"].replace(",", "."))
        estoque_inicial = int(request.form["estoque_inicial"])

        # estoque atual começa igual ao inicial
        estoque_atual = estoque_inicial

        c.execute("""
            INSERT INTO produtos 
            (descricao, valor, estoque_inicial, estoque_atual) 
            VALUES (%s, %s, %s, %s)
        """, (descricao, valor, estoque_inicial, estoque_atual))

        conn.commit()
        flash("Produto cadastrado com sucesso!", "success")

    c.execute("SELECT * FROM produtos ORDER BY descricao ASC")
    lista = c.fetchall()
    conn.close()

    return render_template("produtos.html", produtos=lista)

# =========================
# EDITAR PRODUTOS
# =========================
@app.route("/editar_produto/<int:id>", methods=["GET", "POST"])
def editar_produto(id):
    if session.get("perfil") != "administrador":
        return redirect("/dashboard")

    conn = conectar()
    c = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    if request.method == "POST":
        descricao = request.form["descricao"]
        valor = float(request.form["valor"].replace(",", "."))
        estoque_inicial = int(request.form["estoque_inicial"])

        # 👇 AQUI ESTÁ O AJUSTE
        estoque_atual = estoque_inicial

        c.execute("""
            UPDATE produtos
            SET descricao = %s,
                valor = %s,
                estoque_inicial = %s,
                estoque_atual = %s
            WHERE id = %s
        """, (descricao, valor, estoque_inicial, estoque_atual, id))

        conn.commit()
        conn.close()

        flash("Produto atualizado com sucesso!", "success")
        return redirect("/produtos")

    c.execute("SELECT * FROM produtos WHERE id=%s", (id,))
    produto = c.fetchone()
    conn.close()

    return render_template("editar_produto.html", produto=produto)

# =========================
# ZERAR ESTOQUE
# =========================
@app.route("/zerar_estoque/<int:id>", methods=["POST"])
def zerar_estoque(id):
    if session.get("perfil") != "administrador":
        return redirect("/dashboard")

    conn = conectar()
    c = conn.cursor()

    # Zera apenas o estoque atual
    c.execute("""
        UPDATE produtos
        SET estoque_atual = 0
        WHERE id = %s
    """, (id,))

    conn.commit()
    conn.close()

    flash("Estoque zerado com sucesso!", "warning")
    return redirect("/produtos")

# =========================
# VENDAS (NOVO MODELO)
# =========================

@app.route("/vendas")
def vendas():
    if "usuario" not in session:
        return redirect("/")

    conn = conectar()
    c = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    c.execute("""
        SELECT *
        FROM produtos
        WHERE estoque_atual > 0
        ORDER BY descricao ASC
    """)
    produtos = c.fetchall()
    conn.close()

    return render_template("vendas.html", produtos=produtos)

# =========================
# SALVAR VENDA
# =========================
@app.route("/salvar_venda", methods=["POST"])
def salvar_venda():

    usuario_id = session.get("usuario_id")

    if not usuario_id:
        return jsonify({"erro": "Sessão expirada"}), 403

    dados = request.get_json()

    itens = dados.get("itens", [])
    forma_pagamento = dados.get("forma_pagamento")
    valor_recebido = dados.get("valor_recebido")
    troco = dados.get("troco")

    if not itens:
        return jsonify({"erro": "Nenhum item na venda"}), 400

    if not forma_pagamento:
        return jsonify({"erro": "Forma de pagamento obrigatória"}), 400

    conn = conectar()
    c = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    try:

        # GERA NUMERO DA VENDA (SEQUENCE = seguro para múltiplos caixas)
        c.execute("SELECT nextval('seq_numero_venda') AS numero")
        numero_venda = c.fetchone()["numero"]

        contagem = {}
        for item in itens:
            contagem[item["id"]] = contagem.get(item["id"], 0) + 1

        alertas = []
        venda_registro = []

        # CONTROLE DE ESTOQUE
        for produto_id, quantidade in contagem.items():

            c.execute("""
                UPDATE produtos
                SET estoque_atual = estoque_atual - %s
                WHERE id = %s
                AND estoque_atual >= %s
            """, (quantidade, produto_id, quantidade))

            if c.rowcount == 0:
                conn.rollback()
                return jsonify({"erro": "Estoque insuficiente"}), 400

            c.execute("""
                SELECT descricao, estoque_atual,
                       COALESCE(estoque_minimo,5) as estoque_minimo
                FROM produtos
                WHERE id = %s
            """, (produto_id,))

            produto = c.fetchone()

            venda_registro.append({
                "descricao": produto["descricao"],
                "quantidade": quantidade
            })

            if produto["estoque_atual"] <= produto["estoque_minimo"]:
                alertas.append(
                    f'{produto["descricao"]} com estoque baixo ({produto["estoque_atual"]})'
                )

        # INSERIR ITENS DA VENDA
        for i, item in enumerate(itens):

            c.execute("""
                INSERT INTO vendas
                (numero_venda, produto_id, quantidade, valor_total,
                 data_venda, forma_pagamento, valor_recebido, troco, usuario_id)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, (
                numero_venda,
                item["id"],
                1,
                item["valor"],
                agora_amazonas(),
                forma_pagamento,
                valor_recebido if i == 0 else None,
                troco if i == 0 else None,
                usuario_id
            ))

        conn.commit()

        return jsonify({
            "sucesso": True,
            "numero_venda": numero_venda,
            "data_venda": agora_amazonas().strftime("%d/%m/%Y %H:%M:%S"),
            "forma_pagamento": forma_pagamento,
            "valor_recebido": valor_recebido,
            "troco": troco,
            "alertas": alertas,
            "registro": venda_registro
        })

    except Exception as e:
        conn.rollback()
        return jsonify({"erro": str(e)}), 500

    finally:
        conn.close()
# =========================
# CANCELAR VENDA
# =========================
@app.route("/cancelar_venda", methods=["POST"])
def cancelar_venda():

    if "usuario" not in session:
        return jsonify({"erro": "Não autorizado"}), 403

    dados = request.get_json()
    numero_venda = dados.get("numero_venda")

    if not numero_venda:
        return jsonify({"erro": "Número da venda não informado"}), 400

    conn = conectar()
    c = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    try:
        
        # 🔎 Buscar itens da venda
        c.execute("""
            SELECT produto_id, quantidade
            FROM vendas
            WHERE numero_venda = %s
        """, (numero_venda,))

        itens = c.fetchall()

        if not itens:
            conn.rollback()
            return jsonify({"erro": "Venda não encontrada"}), 404

        # 🔥 Restaurar estoque
        for item in itens:
            c.execute("""
                UPDATE produtos
                SET estoque_atual = estoque_atual + %s
                WHERE id = %s
            """, (item["quantidade"], item["produto_id"]))

        # ❌ Excluir venda
        c.execute("""
            DELETE FROM vendas
            WHERE numero_venda = %s
        """, (numero_venda,))

        conn.commit()

        return jsonify({"sucesso": True})

    except Exception as e:
        conn.rollback()
        return jsonify({"erro": str(e)}), 500

    finally:
        conn.close()

# =========================
# ESTOQUE ATUAL
# =========================
@app.route('/estoque_atual')
def estoque_atual():
    conn = conectar()
    cur = conn.cursor()
    cur.execute("SELECT id, estoque_atual FROM produtos")
    dados = cur.fetchall()
    cur.close()
    conn.close()

    return jsonify(dados)

# =========================
# DASHBOARD AVANÇADO
# =========================
@app.route("/dashboard_avancado")
def dashboard_avancado():
    if "usuario" not in session:
        return redirect("/")

    conn = conectar()
    c = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    # Total geral
    c.execute("SELECT COALESCE(SUM(valor_total),0) as total FROM vendas")
    total_geral = c.fetchone()["total"]

    # Total por forma
    c.execute("""
        SELECT forma_pagamento,
               SUM(valor_total) as total
        FROM vendas
        GROUP BY forma_pagamento
    """)
    por_forma = c.fetchall()

    # Produtos mais vendidos
    c.execute("""
        SELECT p.descricao,
               SUM(v.quantidade) as quantidade,
               SUM(v.valor_total) as total
        FROM vendas v
        JOIN produtos p ON v.produto_id = p.id
        GROUP BY p.descricao
        ORDER BY quantidade DESC
    """)
    mais_vendidos = c.fetchall()

    # Vendas por operador
    c.execute("""
        SELECT u.nome_usuario,
               COUNT(DISTINCT v.numero_venda) AS vendas,
               SUM(v.valor_total) AS total
        FROM vendas v
        JOIN usuarios u ON u.id = v.usuario_id
        GROUP BY u.nome_usuario
        ORDER BY total DESC
    """)
    por_operador = c.fetchall()

    conn.close()

    return render_template(
        "dashboard_avancado.html",
        total_geral=total_geral,
        por_forma=por_forma,
        mais_vendidos=mais_vendidos,
        por_operador=por_operador
    )

# =========================
# RELATÓRIO RESUMO GERAL
# =========================
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import inch

@app.route("/dashboard_avancado_pdf")
def dashboard_avancado_pdf():

    if session.get("perfil") != "administrador":
        return redirect("/dashboard")

    conn = conectar()
    c = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    # 🔎 CONSULTAS
    c.execute("SELECT COALESCE(SUM(valor_total),0) as total FROM vendas")
    total_geral = c.fetchone()["total"]

    c.execute("""
        SELECT forma_pagamento,
               SUM(valor_total) as total
        FROM vendas
        GROUP BY forma_pagamento
    """)
    por_forma = c.fetchall()

    c.execute("""
        SELECT p.descricao,
               SUM(v.quantidade) as quantidade,
               SUM(v.valor_total) as total
        FROM vendas v
        JOIN produtos p ON v.produto_id = p.id
        GROUP BY p.descricao
        ORDER BY quantidade DESC
    """)
    mais_vendidos = c.fetchall()

    c.execute("""
        SELECT u.nome_usuario,
               COUNT(DISTINCT v.numero_venda) AS vendas,
               SUM(v.valor_total) AS total
        FROM vendas v
        JOIN usuarios u ON u.id = v.usuario_id
        GROUP BY u.nome_usuario
        ORDER BY total DESC
    """)
    por_operador = c.fetchall()

    conn.close()

    # 📄 CRIAÇÃO DO PDF
    file_path = "Resumo_Quermesse.pdf"
    doc = SimpleDocTemplate(file_path, pagesize=A4)
    elements = []

    styles = getSampleStyleSheet()

    # 🎨 Estilos personalizados
    titulo_style = ParagraphStyle(
        'TituloCentralizado',
        parent=styles['Title'],
        alignment=1  # centralizado
    )

    destaque_style = ParagraphStyle(
        'Destaque',
        parent=styles['Heading2'],
        textColor=colors.HexColor("#0d6efd")
    )
    
    # 🏷️ Título
    elements.append(Paragraph("QUERMESSE ONLINE", titulo_style))
    elements.append(Spacer(1, 10))
    elements.append(Paragraph("Resumo Geral de Vendas", styles["Heading2"]))
    elements.append(Spacer(1, 10))

    # 📅 Data
    data_atual = agora_amazonas().strftime("%d/%m/%Y %H:%M:%S")
    elements.append(Paragraph(f"Gerado em: {data_atual}", styles["Normal"]))
    elements.append(Spacer(1, 20))

    # 💰 Total Geral
    elements.append(Paragraph(
        f"Total Geral Arrecadado: R$ {round(total_geral,2)}",
        destaque_style
    ))
    elements.append(Spacer(1, 25))

    # =========================
    # 🟦 FUNÇÃO PARA TABELA BONITA
    # =========================
    def tabela_estilizada(dados):
        tabela = Table(dados, hAlign="LEFT")

        tabela.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#a9abaa")),
            ('TEXTCOLOR',(0,0),(-1,0),colors.white),
            ('ALIGN',(1,1),(-1,-1),'CENTER'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
            ('FONTNAME', (0,0), (-1,-1), 'Helvetica'),
            ('BOTTOMPADDING', (0,0), (-1,0), 8),
        ]))

        return tabela

    # 🟩 Forma de pagamento
    elements.append(Paragraph("Vendas por Forma de Pagamento", styles["Heading3"]))
    elements.append(Spacer(1,10))

    data_forma = [["Forma de Pagamento", "Total (R$)"]]
    for item in por_forma:
        data_forma.append([
            item["forma_pagamento"],
            round(item["total"],2)
        ])

    elements.append(tabela_estilizada(data_forma))
    elements.append(Spacer(1,25))

    # 🟩 Produtos
    elements.append(Paragraph("Vendas por Produtos", styles["Heading3"]))
    elements.append(Spacer(1,10))

    data_prod = [["Produto", "Quantidade", "Total (R$)"]]
    for p in mais_vendidos:
        data_prod.append([
            p["descricao"],
            p["quantidade"],
            round(p["total"],2)
        ])

    elements.append(tabela_estilizada(data_prod))
    elements.append(Spacer(1,25))

    # 🟩 Operadores
    elements.append(Paragraph("Vendas por Operador de Caixa", styles["Heading3"]))
    elements.append(Spacer(1,10))

    data_op = [["Operador", "Nº Vendas", "Total (R$)"]]
    for o in por_operador:
        data_op.append([
            o["nome_usuario"],
            o["vendas"],
            round(o["total"],2)
        ])

    elements.append(tabela_estilizada(data_op))
    elements.append(Spacer(1,30))

    # 📝 Rodapé
    elements.append(Paragraph(
        "Relatório gerado automaticamente pelo Sistema Quermesse Online.",
        styles["Italic"]
    ))

    doc.build(elements)

    return send_file(file_path, as_attachment=True)

# =========================
# FECHAMENTO
# =========================
@app.route("/fechamento")
def fechamento():

    if "usuario" not in session:
        return redirect("/")

    data = request.args.get("data")

    if not data:
        data = agora_amazonas().strftime("%Y-%m-%d")

    conn = conectar()
    c = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    c.execute("""
        SELECT forma_pagamento,
               SUM(valor_total) AS total,
               SUM(COALESCE(troco_unico,0)) AS total_troco
        FROM (
            SELECT numero_venda,
                   forma_pagamento,
                   SUM(valor_total) AS valor_total,
                   MAX(COALESCE(troco,0)) AS troco_unico
            FROM vendas
            WHERE DATE(data_venda) = %s
            GROUP BY numero_venda, forma_pagamento
        ) sub
        GROUP BY forma_pagamento
        ORDER BY forma_pagamento
    """, (data,))

    resultado = c.fetchall()

    conn.close()

    # CALCULAR TOTAIS
    total_geral = sum(float(r["total"] or 0) for r in resultado)
    total_troco = sum(float(r["total_troco"] or 0) for r in resultado)

    return render_template(
        "fechamento.html",
        resultado=resultado,
        data=data,
        total_geral=total_geral,
        total_troco=total_troco
    )
    
# =========================
# FECHAMENTO PDF
# =========================
@app.route("/fechamento_pdf")
def fechamento_pdf():

    if "usuario" not in session:
        return redirect("/")

    data = request.args.get("data")

    if not data:
        data = agora_amazonas().strftime("%Y-%m-%d")

    conn = conectar()
    c = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    c.execute("""
    SELECT forma_pagamento,
           SUM(valor_total) AS total,
           SUM(COALESCE(troco_unico,0)) AS total_troco
    FROM (
        SELECT numero_venda,
               forma_pagamento,
               SUM(valor_total) AS valor_total,
               MAX(COALESCE(troco,0)) AS troco_unico
        FROM vendas
        WHERE DATE(data_venda) = %s
        GROUP BY numero_venda, forma_pagamento
    ) sub
    GROUP BY forma_pagamento
    """, (data,))

    resultado = c.fetchall()
    conn.close()

    # ======================
    # CRIAÇÃO DO PDF
    # ======================

    file_path = "Fechamento_Caixa.pdf"
    doc = SimpleDocTemplate(file_path, pagesize=A4)

    elements = []
    styles = getSampleStyleSheet()

    titulo_style = ParagraphStyle(
        'TituloCentralizado',
        parent=styles['Title'],
        alignment=1
    )

    elements.append(Paragraph("QUERMESSE ONLINE", titulo_style))
    elements.append(Spacer(1, 10))
    elements.append(Paragraph("Fechamento de Caixa", styles["Heading2"]))
    elements.append(Spacer(1, 10))
    elements.append(Paragraph(f"Data: {data}", styles["Normal"]))
    elements.append(Spacer(1, 20))

    tabela = [["Forma Pagamento", "Total Vendido", "Total Troco"]]

    total_vendido = 0
    total_troco = 0

    for r in resultado:
        total = float(r["total"] or 0)
        troco = float(r["total_troco"] or 0)
    
        total_vendido += total
        total_troco += troco
    
        tabela.append([
            r["forma_pagamento"],
            f"R$ {round(total,2)}",
            f"R$ {round(troco,2)}"
        ])
        
    tabela.append([
        "TOTAL GERAL",
        f"R$ {round(total_vendido,2)}",
        f"R$ {round(total_troco,2)}"
    ])

    table = Table(tabela)

    table.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,0),'#a9abaa'),
        ('TEXTCOLOR',(0,0),(-1,0),'white'),
        ('GRID',(0,0),(-1,-1),0.5,'grey'),
        ('FONTNAME',(0,0),(-1,-1),'Helvetica'),
        ('BACKGROUND', (0,-1), (-1,-1), '#d9edf7')
    ]))

    elements.append(table)

    elements.append(Spacer(1,20))
    elements.append(Paragraph(
        "Relatório gerado automaticamente pelo Sistema Quermesse Online.",
        styles["Italic"]
    ))

    doc.build(elements)

    return send_file(file_path, as_attachment=True)

# =========================
# RELATÓRIOS
# =========================
@app.route("/relatorios")
def relatorios():

    if "usuario" not in session:
        return redirect("/")

    if session.get("perfil") != "administrador":
        return redirect("/dashboard")

    data_inicio = request.args.get("data_inicio")
    data_fim = request.args.get("data_fim")
    forma_pagamento = request.args.get("forma_pagamento")
    usuario_id = request.args.get("usuario_id")
    numero_venda = request.args.get("numero_venda")
    
    # corrigir parâmetros vindos como "None"
    if data_inicio == "None":
        data_inicio = None
    
    if data_fim == "None":
        data_fim = None
    
    if forma_pagamento == "None":
        forma_pagamento = None
    
    if usuario_id == "None":
        usuario_id = None

    if numero_venda == "None":
        numero_venda = None
    
    conn = conectar()
    c = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    query = """
        SELECT 
            v.numero_venda,
            MIN(v.data_venda AT TIME ZONE 'America/Manaus') AS data_venda,
            v.forma_pagamento,
            u.nome_usuario,
            SUM(v.valor_total) AS total
        FROM vendas v
        JOIN usuarios u ON u.id = v.usuario_id
        WHERE 1=1
    """

    params = []

    if data_inicio and data_inicio != "None":
        query += " AND DATE(v.data_venda) >= %s"
        params.append(data_inicio)

    if data_fim and data_fim != "None":
        query += " AND DATE(v.data_venda) <= %s"
        params.append(data_fim)

    if forma_pagamento:
        query += " AND v.forma_pagamento = %s"
        params.append(forma_pagamento)

    if usuario_id:
        query += " AND u.id = %s"
        params.append(usuario_id)

    if numero_venda and numero_venda != "None":
        query += " AND v.numero_venda = %s"
        params.append(numero_venda)

    query += """
        GROUP BY 
            v.numero_venda,
            v.forma_pagamento,
            u.nome_usuario
        ORDER BY v.numero_venda DESC
    """

    c.execute(query, params)
    vendas = c.fetchall()

    # usuários para filtro
    c.execute("SELECT id, nome_usuario FROM usuarios ORDER BY nome_usuario")
    usuarios = c.fetchall()

    # formas de pagamento
    c.execute("SELECT DISTINCT forma_pagamento FROM vendas ORDER BY forma_pagamento")
    formas = c.fetchall()

    conn.close()

    total_geral = sum(float(v["total"] or 0) for v in vendas)

    return render_template(
        "relatorios.html",
        vendas=vendas,
        total_geral=total_geral,
        usuarios=usuarios,
        formas=formas,
        filtro_data_inicio=data_inicio,
        filtro_data_fim=data_fim,
        filtro_forma_pagamento=forma_pagamento,
        filtro_usuario_id=usuario_id,
        filtro_numero_venda=numero_venda
    )
    
# =========================
# PDF
# =========================
@app.route("/relatorio_vendas_pdf")
def relatorio_vendas_pdf():

    if "usuario" not in session:
        return redirect("/")

    data_inicio = request.args.get("data_inicio")
    data_fim = request.args.get("data_fim")
    forma_pagamento = request.args.get("forma_pagamento")
    usuario_id = request.args.get("usuario_id")
    numero_venda = request.args.get("numero_venda")

    conn = conectar()
    c = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    query = """
        SELECT 
            v.numero_venda,
            MIN(v.data_venda AT TIME ZONE 'America/Manaus') as data_venda,
            SUM(v.valor_total) as valor_total,
            v.forma_pagamento,
            u.nome_usuario
        FROM vendas v
        JOIN usuarios u ON u.id = v.usuario_id
        WHERE 1=1
    """

    params = []

    if data_inicio:
        query += " AND DATE(v.data_venda) >= %s"
        params.append(data_inicio)

    if data_fim:
        query += " AND DATE(v.data_venda) <= %s"
        params.append(data_fim)

    if forma_pagamento:
        query += " AND v.forma_pagamento = %s"
        params.append(forma_pagamento)

    if usuario_id:
        query += " AND v.usuario_id = %s"
        params.append(usuario_id)

    if numero_venda:
        query += " AND v.numero_venda = %s"
        params.append(numero_venda) 

    query += """
        GROUP BY v.numero_venda, v.forma_pagamento, u.nome_usuario
        ORDER BY v.numero_venda DESC
    """

    c.execute(query, params)
    vendas = c.fetchall()

    # ===== GERAR PDF =====
    caminho = "relatorio_vendas.pdf"
    
    elementos = []
    
    styles = getSampleStyleSheet()
    
    # Título
    titulo = Paragraph("Relatório de Vendas", styles["Title"])
    elementos.append(titulo)
    elementos.append(Spacer(1,20))
    
    dados = [["Venda", "Data", "Pagamento", "Usuário", "Valor"]]
    
    total = 0
    
    for v in vendas:
        dados.append([
            v["numero_venda"],
            v["data_venda"].strftime("%d/%m/%Y %H:%M"),
            v["forma_pagamento"],
            v["nome_usuario"],
            f"R$ {v['valor_total']:.2f}"
        ])
        total += v["valor_total"]
    
    dados.append(["", "", "", "TOTAL", f"R$ {total:.2f}"])
    
    tabela = Table(dados, colWidths=[60,120,110,120,80])
    
    tabela.setStyle(TableStyle([
    
        # Cabeçalho
        ('BACKGROUND',(0,0),(-1,0),colors.lightgrey),
        ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
        ('ALIGN',(0,0),(-1,0),'CENTER'),
    
        # Alinhamento das colunas
        ('ALIGN',(0,1),(0,-1),'CENTER'),
        ('ALIGN',(1,1),(3,-1),'CENTER'),
        ('ALIGN',(4,1),(4,-1),'RIGHT'),
    
        # Grade
        ('GRID',(0,0),(-1,-1),1,colors.black),
    
        # Linha total
        ('FONTNAME',(3,-1),(4,-1),'Helvetica-Bold'),
        ('BACKGROUND',(3,-1),(4,-1),colors.lightgrey)
    
    ]))
    
    elementos.append(tabela)
    
    doc = SimpleDocTemplate(caminho, pagesize=A4)
    doc.build(elementos)

    conn.close()

    return send_file(caminho, as_attachment=True)
# =========================
# EXCEL
# =========================
@app.route("/relatorio_vendas_excel")
def relatorio_vendas_excel():

    if session.get("perfil") != "administrador":
        return redirect("/dashboard")

    # filtros vindos da URL
    data_inicio = request.args.get("data_inicio")
    data_fim = request.args.get("data_fim")
    forma_pagamento = request.args.get("forma_pagamento")
    usuario_id = request.args.get("usuario_id")
    numero_venda = request.args.get("numero_venda")
    
    # Corrigir valores inválidos vindos da URL
    if data_inicio in ("", "None"):
        data_inicio = None
    
    if data_fim in ("", "None"):
        data_fim = None
    
    if forma_pagamento in ("", "None"):
        forma_pagamento = None
    
    if usuario_id in ("", "None"):
        usuario_id = None

    if numero_venda in ("", "None"):
        numero_venda = None

    conn = conectar()
    c = conn.cursor()

    sql = """
        SELECT 
            v.numero_venda,
            MIN(v.data_venda AT TIME ZONE 'America/Manaus') AS data_venda,
            v.forma_pagamento,
            u.nome_usuario,
            SUM(v.valor_total) AS total
        FROM vendas v
        JOIN usuarios u ON u.id = v.usuario_id
        WHERE 1=1
    """

    params = []

    if data_inicio:
        sql += " AND DATE(v.data_venda) >= %s"
        params.append(data_inicio)

    if data_fim:
        sql += " AND DATE(v.data_venda) <= %s"
        params.append(data_fim)

    if forma_pagamento:
        sql += " AND v.forma_pagamento = %s"
        params.append(forma_pagamento)

    if usuario_id:
        sql += " AND v.usuario_id = %s"
        params.append(usuario_id)

    if numero_venda:
        sql += " AND v.numero_venda = %s"
        params.append(numero_venda)

    sql += """
        GROUP BY 
            v.numero_venda, 
            v.forma_pagamento, 
            u.nome_usuario
        ORDER BY v.numero_venda DESC
    """

    c.execute(sql, params)

    vendas = c.fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active

    # Cabeçalho
    headers = ["ID Venda", "Data", "Forma Pagamento", "Valor Total", "Usuário"]
    ws.append(headers)

    # ===== ESTILO DO CABEÇALHO =====
    bold_font = Font(bold=True)
    fill_gray = PatternFill(start_color="DDDDDD",
                            end_color="DDDDDD",
                            fill_type="solid")

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = bold_font
        cell.fill = fill_gray

    # Inserindo dados
    total_geral = 0
    
    for v in vendas:
        valor = float(v[4])
        total_geral += valor

        ws.append([
            v[0],
            v[1].strftime("%d/%m/%Y %H:%M:%S"),
            v[2],
            valor,
            v[3]
        ])

    # ===== LINHA DO TOTAL =====
    linha_total = ws.max_row + 1
    
    # Texto
    ws.cell(row=linha_total, column=1).value = "Total das vendas"
    
    # Valor
    ws.cell(row=linha_total, column=4).value = total_geral
    
    # Mesclar A+B+C
    ws.merge_cells(start_row=linha_total, start_column=1, end_row=linha_total, end_column=3)
    
    # Mesclar D+E
    ws.merge_cells(start_row=linha_total, start_column=4, end_row=linha_total, end_column=5)
    
    # Negrito
    ws.cell(row=linha_total, column=1).font = Font(bold=True)
    ws.cell(row=linha_total, column=4).font = Font(bold=True)

    ws.cell(row=linha_total, column=1).alignment = Alignment(horizontal="right")
    ws.cell(row=linha_total, column=4).alignment = Alignment(horizontal="center")

    # ===== AJUSTE AUTOMÁTICO DE LARGURA =====
    for column_cells in ws.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter

        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        adjusted_width = max_length + 2
        ws.column_dimensions[column_letter].width = adjusted_width

    file_path = "Relatorio_Vendas.xlsx"
    wb.save(file_path)

    return send_file(file_path, as_attachment=True)
# =========================
# RESETAR QUERMESSE
# =========================
@app.route('/resetar_quermesse', methods=['POST'])
def resetar_quermesse():

    if "usuario" not in session:
        return redirect("/")

    # 👑 Verificação correta
    if session.get("perfil") != "administrador":
        flash("Acesso restrito!", "danger")
        return redirect("/dashboard")

    conn = conectar()
    if not conn:
        flash("Erro ao conectar no banco!", "danger")
        return redirect("/dashboard")

    cur = conn.cursor()

    try:
        # 🔥 Apaga TODAS as vendas
        cur.execute("TRUNCATE TABLE vendas RESTART IDENTITY CASCADE;")

        # 🔄 Opcional: restaurar estoque para inicial
        cur.execute("""
            UPDATE produtos
            SET estoque_atual = estoque_inicial
        """)

        conn.commit()
        flash("Sistema resetado para nova quermesse com sucesso!", "success")

    except Exception as e:
        conn.rollback()
        print("ERRO RESET:", e)
        flash("Erro ao resetar sistema!", "danger")

    finally:
        conn.close()

    return redirect("/dashboard")

@app.route("/health")
def health():
    return "OK", 200

